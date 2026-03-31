from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import pandas as pd

from src.transform import build_current_dataset, build_powerbi_star_schema
from src.validate import normalize_name


STATUS_HISTORY_COLUMNS = [
    "contenedor_id",
    "pedido",
    "naviera",
    "puerto",
    "fecha_cas",
    "plan_llegada_grupasa",
    "plan_devolucion_vacio",
    "deposito_vacio",
    "status_actual",
    "horario_entrega_real",
    "tipo_incidencia",
    "comentario_status",
    "fecha_snapshot",
]


@dataclass(frozen=True)
class InventoryWorkbook:
    path: Path
    snapshot_date: date
    inventory_df: pd.DataFrame


def load_inventory_workbook(path: Path) -> InventoryWorkbook:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "INVENTARIO" not in workbook.sheetnames:
        raise ValueError(f"No existe la hoja INVENTARIO en {path}")

    snapshot_date = _resolve_snapshot_date(path, workbook.properties.modified)
    worksheet = workbook["INVENTARIO"]
    rows = list(worksheet.iter_rows(min_row=1, max_row=2, values_only=True))
    if not rows:
        raise ValueError(f"La hoja INVENTARIO está vacía en {path}")

    raw_headers = [value for value in rows[0]]
    normalized_headers = [_normalize_inventory_header(value) for value in raw_headers]
    data_rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        record = {normalized_headers[index]: row[index] if index < len(row) else None for index in range(len(normalized_headers))}
        data_rows.append(record)

    inventory_df = pd.DataFrame(data_rows)
    if inventory_df.empty:
        inventory_df = pd.DataFrame(columns=normalized_headers)

    inventory_df = _standardize_inventory_dataframe(inventory_df)
    return InventoryWorkbook(path=path, snapshot_date=snapshot_date, inventory_df=inventory_df)


def build_historical_outputs(workbooks: list[InventoryWorkbook], cas_alert_days: int) -> dict[str, pd.DataFrame]:
    if not workbooks:
        raise ValueError("Debes proporcionar al menos un workbook de inventario")

    ordered_workbooks = sorted(workbooks, key=lambda item: item.snapshot_date)
    status_history = synthesize_status_history(ordered_workbooks)
    latest_workbook = ordered_workbooks[-1]
    latest_inventory = latest_workbook.inventory_df.copy()

    registro_df = build_registro_sheet(latest_inventory)
    plan_grupasa_df = build_plan_grupasa_sheet(latest_inventory)
    plan_galagans_df = build_plan_galagans_sheet(latest_inventory)
    status_actual_df = build_status_actual_sheet(latest_inventory, latest_workbook.snapshot_date)

    sheets = {
        "Registro_Contenedores": registro_df,
        "Planif_Grupasa": plan_grupasa_df,
        "Planif_Galagans": plan_galagans_df,
        "Status_Operativo": status_actual_df,
    }

    grupasa_plan_resolved = plan_grupasa_df.assign(plan_slot=pd.NA, tipo_asignacion="directa_hoja")
    current_dataset = build_current_dataset(
        sheets=sheets,
        status_history=status_history,
        grupasa_plan_resolved=grupasa_plan_resolved,
        snapshot_date=latest_workbook.snapshot_date.isoformat(),
        cas_alert_days=cas_alert_days,
    )
    powerbi_outputs = build_powerbi_star_schema(
        current_dataset=current_dataset,
        status_history=status_history,
        cas_alert_days=cas_alert_days,
    )

    return {
        "status_historico": status_history,
        "registro_congelado": registro_df.assign(fecha_snapshot=latest_workbook.snapshot_date.isoformat()),
        "plan_galagans_congelado": plan_galagans_df.assign(fecha_snapshot=latest_workbook.snapshot_date.isoformat()),
        "contenedores_actual": current_dataset,
        **powerbi_outputs,
    }


def synthesize_status_history(workbooks: list[InventoryWorkbook]) -> pd.DataFrame:
    history_frames: list[pd.DataFrame] = []
    for workbook in workbooks:
        frame = _build_status_history_for_inventory(workbook.inventory_df, workbook.snapshot_date)
        frame["source_snapshot_date"] = workbook.snapshot_date.isoformat()
        history_frames.append(frame)

    if not history_frames:
        return pd.DataFrame(columns=STATUS_HISTORY_COLUMNS)

    combined = pd.concat(history_frames, ignore_index=True, sort=False)
    combined["fecha_snapshot"] = pd.to_datetime(combined["fecha_snapshot"], errors="coerce").dt.date
    combined["source_snapshot_date"] = pd.to_datetime(combined["source_snapshot_date"], errors="coerce").dt.date
    combined = combined.sort_values(["fecha_snapshot", "contenedor_id", "source_snapshot_date"])
    combined = combined.drop_duplicates(subset=["fecha_snapshot", "contenedor_id"], keep="last")
    combined = combined.drop(columns=["source_snapshot_date"])
    combined["fecha_snapshot"] = combined["fecha_snapshot"].astype("string")
    return combined[STATUS_HISTORY_COLUMNS].reset_index(drop=True)


def build_registro_sheet(inventory_df: pd.DataFrame) -> pd.DataFrame:
    return inventory_df[
        [
            "contenedor_id",
            "pedido",
            "parcial",
            "naviera",
            "puerto",
            "deposito_vacio",
            "fecha_arribo_gye",
            "fecha_salida_autorizada",
            "fecha_retiro_puerto",
            "fecha_cas",
        ]
    ].rename(columns={"fecha_retiro_puerto": "fecha_arribo"}).reset_index(drop=True)


def build_plan_grupasa_sheet(inventory_df: pd.DataFrame) -> pd.DataFrame:
    return inventory_df[
        [
            "contenedor_id",
            "pedido",
            "naviera",
            "puerto",
            "deposito_vacio",
            "fecha_cas",
            "plan_llegada_grupasa",
            "bodega",
        ]
    ].assign(hora_descarga=pd.NA, comentario_plan_grupasa=pd.NA).reset_index(drop=True)


def build_plan_galagans_sheet(inventory_df: pd.DataFrame) -> pd.DataFrame:
    return inventory_df[
        [
            "contenedor_id",
            "pedido",
            "naviera",
            "puerto",
            "deposito_vacio",
            "plan_llegada_patio",
            "plan_devolucion_vacio",
        ]
    ].assign(comentario_plan_galagans=pd.NA).reset_index(drop=True)


def build_status_actual_sheet(inventory_df: pd.DataFrame, snapshot_date: date) -> pd.DataFrame:
    frame = inventory_df[
        [
            "contenedor_id",
            "pedido",
            "naviera",
            "puerto",
            "fecha_cas",
            "plan_llegada_grupasa",
            "plan_devolucion_vacio",
            "deposito_vacio",
            "tipo_incidencia",
        ]
    ].copy()
    frame["status_actual"] = inventory_df.apply(lambda row: _status_for_date(row, snapshot_date), axis=1)
    frame["horario_entrega_real"] = pd.NA
    frame["comentario_status"] = "Inferido desde inventario historico"
    return frame.reset_index(drop=True)


def _build_status_history_for_inventory(inventory_df: pd.DataFrame, snapshot_date: date) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    for row in inventory_df.to_dict(orient="records"):
        base = {
            "contenedor_id": row.get("contenedor_id"),
            "pedido": row.get("pedido"),
            "naviera": row.get("naviera"),
            "puerto": row.get("puerto"),
            "fecha_cas": _to_iso(row.get("fecha_cas")),
            "plan_llegada_grupasa": _to_iso(row.get("plan_llegada_grupasa")),
            "plan_devolucion_vacio": _to_iso(row.get("plan_devolucion_vacio")),
            "deposito_vacio": row.get("deposito_vacio"),
            "horario_entrega_real": pd.NA,
            "tipo_incidencia": row.get("tipo_incidencia"),
            "comentario_status": "Inferido desde inventario historico",
        }
        for current_date, status in _iter_status_days(row, snapshot_date):
            records.append(
                {
                    **base,
                    "status_actual": status,
                    "fecha_snapshot": current_date.isoformat(),
                }
            )

    if not records:
        return pd.DataFrame(columns=STATUS_HISTORY_COLUMNS)
    return pd.DataFrame(records)


def _iter_status_days(row: dict[str, Any], snapshot_date: date) -> list[tuple[date, str]]:
    puerto_start = _first_date(row.get("fecha_arribo_gye"), row.get("fecha_salida_autorizada"), row.get("fecha_retiro_puerto"))
    patio_start = _as_date(row.get("fecha_retiro_puerto"))
    bodega_start = _first_date(row.get("fecha_traslado_planta"))
    deposito_start = _as_date(row.get("fecha_devolucion"))

    segments: list[tuple[date | None, date | None, str]] = [
        (puerto_start, _day_before(patio_start) if patio_start else snapshot_date, "EN PUERTO"),
        (
            patio_start,
            _day_before(bodega_start) if bodega_start else (_day_before(deposito_start) if deposito_start else snapshot_date),
            "EN PATIO",
        ),
        (bodega_start, _day_before(deposito_start) if deposito_start else snapshot_date, "EN GRUPASA"),
        (deposito_start, snapshot_date, "DEVUELTO DEPOSITO VACIO"),
    ]

    results: list[tuple[date, str]] = []
    for start_date, end_date, status in segments:
        if start_date is None or end_date is None:
            continue
        if start_date > snapshot_date:
            continue
        end_date = min(end_date, snapshot_date)
        if start_date > end_date:
            continue
        current_date = start_date
        while current_date <= end_date:
            results.append((current_date, status))
            current_date += timedelta(days=1)
    return results


def _standardize_inventory_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    rename_map = {
        "pedido": "pedido",
        "parcial": "parcial",
        "parcial_": "parcial",
        "puerto": "puerto",
        "bl": "bl",
        "naviera": "naviera",
        "contenedor": "contenedor_id",
        "tipo_de_contenedor": "tipo_contenedor",
        "producto": "producto",
        "fecha_de_arribo_a_gye": "fecha_arribo_gye",
        "fecha_de_salida_autorizada": "fecha_salida_autorizada",
        "fecha_de_retiro_de_puerto": "fecha_retiro_puerto",
        "cas": "fecha_cas",
        "fecha_descarga_planificada": "plan_llegada_grupasa",
        "fecha_descarga_planificada_": "plan_llegada_grupasa",
        "fecha_de_traslado_a_planta": "fecha_traslado_planta",
        "bodega_de_traslado": "bodega",
        "deposito_de_devolucion": "deposito_vacio",
        "fecha_de_devolucion": "fecha_devolucion",
        "observaciones": "comentario",
    }

    normalized_columns = {column: rename_map.get(column, column) for column in df.columns}
    standardized = df.rename(columns=normalized_columns).copy()
    standardized = standardized.loc[:, ~standardized.columns.duplicated()]

    for column in [
        "contenedor_id",
        "pedido",
        "parcial",
        "naviera",
        "puerto",
        "deposito_vacio",
        "bodega",
        "comentario",
    ]:
        if column in standardized.columns:
            standardized[column] = standardized[column].astype("string").str.strip()

    for column in [
        "fecha_arribo_gye",
        "fecha_salida_autorizada",
        "fecha_retiro_puerto",
        "fecha_cas",
        "plan_llegada_grupasa",
        "fecha_traslado_planta",
        "fecha_devolucion",
    ]:
        if column in standardized.columns:
            standardized[column] = standardized[column].apply(_as_date)

    standardized["plan_llegada_patio"] = standardized.get("fecha_retiro_puerto")
    standardized["plan_devolucion_vacio"] = standardized.get("fecha_devolucion")
    standardized["tipo_incidencia"] = standardized.get("tipo_incidencia", pd.Series(["SIN NOVEDAD"] * len(standardized))).fillna("SIN NOVEDAD")
    standardized = standardized.dropna(subset=["contenedor_id"])
    standardized = standardized.drop_duplicates(subset=["contenedor_id"], keep="last")
    return standardized.reset_index(drop=True)


def _normalize_inventory_header(value: object) -> str:
    text = normalize_name(value)
    if text.startswith("n_") or text == "n":
        return "numero"
    return text


def _resolve_snapshot_date(path: Path, modified_value: datetime | None) -> date:
    if modified_value is not None:
        return modified_value.date()
    return datetime.fromtimestamp(path.stat().st_mtime).date()


def _status_for_date(row: pd.Series | dict[str, Any], snapshot_date: date) -> str:
    deposito_start = _as_date(row.get("fecha_devolucion"))
    bodega_start = _as_date(row.get("fecha_traslado_planta"))
    patio_start = _as_date(row.get("fecha_retiro_puerto"))
    puerto_start = _first_date(row.get("fecha_arribo_gye"), row.get("fecha_salida_autorizada"), row.get("fecha_retiro_puerto"))

    if deposito_start and deposito_start <= snapshot_date:
        return "DEVUELTO DEPOSITO VACIO"
    if bodega_start and bodega_start <= snapshot_date:
        return "EN GRUPASA"
    if patio_start and patio_start <= snapshot_date:
        return "EN PATIO"
    if puerto_start and puerto_start <= snapshot_date:
        return "EN PUERTO"
    return "SIN_STATUS"


def _first_date(*values: object) -> date | None:
    parsed = [_as_date(value) for value in values]
    valid = [value for value in parsed if value is not None]
    return min(valid) if valid else None


def _day_before(value: date | None) -> date | None:
    if value is None:
        return None
    return value - timedelta(days=1)


def _as_date(value: object) -> date | None:
    if value is None or value == "" or pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    converted = pd.to_datetime(value, errors="coerce")
    if pd.isna(converted):
        return None
    return converted.date()


def _to_iso(value: object) -> str | pd.NA:
    parsed = _as_date(value)
    return parsed.isoformat() if parsed else pd.NA
